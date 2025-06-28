import os
import time
import logging
from decimal import Decimal
from urllib.parse import urlencode, quote
import aiohttp
from aiogram.filters import callback_data
from aiogram.fsm.context import FSMContext
from openpyxl import load_workbook

from aiogram.types import CallbackQuery, InlineKeyboardButton, InlineKeyboardMarkup
from aiogram import Router, F
from aiogram.fsm.state import State, StatesGroup
from dotenv import load_dotenv

from src.base.database import DATABASE




router = Router()
logger = logging.getLogger(__name__)





class SetState(StatesGroup):
    sum = State()

class sum:
    sum = 20

SUM = sum()

dotenv_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))), 'run', 'config.env')
print(f"Dotenv path: {dotenv_path} - exists: {os.path.exists(dotenv_path)}")
load_dotenv(dotenv_path)

yoomoney_access_token = str(os.getenv('yoomoney_access_token'))
yoomoney_receiver = str(os.getenv('yoomoney_receiver'))
yoomoney_scopes = str(os.getenv('yoomoney_scopes'))
yoomoney_redirect_url = str(os.getenv('yoomoney_redirect_url'))
yoomoney_client_id = str(os.getenv('yoomoney_client_id'))
debug_yoomoney_mode = str(os.getenv('debug_yoomoney_mode'))

print(yoomoney_access_token)
print(yoomoney_receiver)
print(yoomoney_scopes)
print(yoomoney_redirect_url)
print(yoomoney_client_id)





def create_quickpay_link(amount, user_id):
    label = f"payment_{user_id}_{amount:.2f}_{int(time.time())}"
    params = {
        "receiver": yoomoney_receiver,
        "quickpay-form": "shop",
        "targets": f"Пополнение {amount}",
        "paymentType": "AC",
        "sum": f"{amount:.2f}",
        "label": label,
        "successURL": ""
    }
    encoded = urlencode(params, quote_via=quote)
    url = f"https://yoomoney.ru/quickpay/confirm.xml?{encoded}"
    print(url, label)
    print("yoomoney_scopes =", yoomoney_scopes)

    return url, label


@router.callback_query(F.data.startswith('sub_check_'))
async def check_yoomoney_payment(callback: CallbackQuery):
    data = callback.data
    parts = data.split('_')
    user_id = int(parts[3])
    print(user_id)
    label = '_'.join(parts[1:])

    if debug_yoomoney_mode:
        print("[DEMO] YooMoney check is always successful for label:", label)
        await add_excel(user_id)
        await callback.answer("Оплата подтверждена (DEMO режим)!")
        await callback.message.edit_text("Оплата подтверждена!\nСпасибо за покупку!\nЗаказ можно посмотреть в корзине")
        return True

    headers = {
        "Authorization": f"Bearer {yoomoney_access_token}",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    payload = {
        "label": label,
        "records": 10,
        "type": "deposition"
    }
    async with aiohttp.ClientSession() as session:
        try:
            async with session.post(yoomoney_scopes, headers=headers, data=payload) as resp:
                if resp.status == 200:
                    result = await resp.json()
                    for op in result.get("operations", []):
                        if op.get("label") == label and op.get("status") == "success":
                            await add_excel(user_id)
                            await callback.answer("Оплата подтверждена!")
                            await callback.message.edit_text("Оплата подтверждена!\nСпасибо за покупку!\nЗаказ можно посмотреть в корзине")
                            return STrue
                else:
                    logger.error(f"YooMoney API error: {resp.status} {await resp.text()}")
        except Exception as e:
            logger.exception(f"Error while checking payment in YooMoney: {e}")
    await callback.answer("Ошибка при проверке оплаты.")
    
    
async def add_excel(user_id, get = None):
    user_info_list = await DATABASE.get_cb_for_id(user_id, "users")
    if not user_info_list:
        print("Пользователь не найден")
        return False
    user_info = user_info_list[0]
    
    product_in_order = await DATABASE.get_cb_for_id(user_id, "basket")
    text = ""
    summa = 0
    
    for i, item in enumerate(product_in_order):
        if int(item['pay']) == 0:
            product_id = item['product']
            quantity = item['quantity']
            product_info = await DATABASE.get_cb_for_id(product_id, "catalog")
            prod = product_info[0]
            name = prod['name']
            price = prod['price']
            partial = int(price) * int(quantity)
            text += f"{i + 1}. {name} - {quantity} шт. - {partial} руб.\n\n"
            summa += partial
    if get is not None:
        text = f"{user_info['address']}\n"
        text += f"{user_info['phone']}\n"
        text += f"{user_info['name']}\n"
        text += f"{user_info['lastname']}\n"
        text += f"{user_info['email']}\n"
        return text
    
    PROJECT_ROOT = os.path.dirname(
        os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
    excel_file_path = os.path.join(PROJECT_ROOT, 'database.xlsx')
    print(excel_file_path)
    fn = excel_file_path
    
    wb = load_workbook(fn)

    if 'orders' not in wb.sheetnames:
        wb.create_sheet('orders')
    ws = wb['orders']
    print(user_info)
    ws.append([
        f"{user_info['id']}",
        f"{user_info['role']}",
        f"{user_info['address']}",
        f"{user_info['phone']}",
        f"{user_info['name']}",
        f"{user_info['lastname']}",
        f"{user_info['email']}",
        f"{text}",
        f"{summa}"
    ])
    wb.save(fn)
    wb.close()
    return True

